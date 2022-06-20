from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from .regexes  import digit_regex
from fuzzywuzzy import fuzz
from .packages.transliterate import to_cyrillic, to_latin
from config.settings import MEDIA_ROOT


TYPES = (
    "супп", "таб","р-р", "инф.", "саше",
    "капс", 'гел', "лосьон",
    "сироп", "масло", "сусп", "пор", 
    "спрей", "шампун", "пастилки",
    "шип", "лиоф", "капли", "душ", "крем",
    "паст", "маз", "инсулин",
    "рект", "гранулы", "инфузия", "жидкий",
    "увл", "sprey", 'аэр', 

    "supp", "tab","r-r", "inf.", "sashe",
    "kaps", 'gel', "los'on", 'losyon',
    'lasyon', "sirop", "maslo", "susp", "por", 
    "shampun", "pastilki",
    "ship", "liof", "kapli", "dush", "krem",
    "past", "maz", "insulin",
    "rekt", "granuly", "infuziya", "jidkiy",
    "uvl", 'aer.'
    )

EXTRA_TYPES = (
   "ультра", 'ultra',  "супер", "плюс", "мини", "д/детей", "детский", 
    "plus", "super", "baby", "беби"
)

MEASURES = ("гр", "мг", "г", "мл", "МЕ","мкг", "ед", "%", "xxl", "M", "S", "2xl", "xl")


def make_comparison(data):
    med_col_1 = int(data.cleaned_data['med_name_col_1'])
    com_col_1 = int(data.cleaned_data['com_name_col_1'])
    file_1 = data.cleaned_data['file_1']

    med_col_2 = int(data.cleaned_data['med_name_col_2'])
    com_col_2 = int(data.cleaned_data['com_name_col_2'])
    file_2 = data.cleaned_data['file_2']

    wb = Workbook()
    first_wb = None
    second_wb = None
    try:
        first_wb = load_workbook(file_1)  #loading a workbook
        second_wb = load_workbook(file_2)
    except Exception as ex:
        return HttpResponse(ex)
    
    
    ws_1 = first_wb.active
    ws_2 = second_wb.active
    
    NEW_FILE_VAlUES = () #ALL VALUES of COLUMNS FOR WRITING DATA

    title_row = ()   #for name and title of cols
    for row in ws_1.iter_rows(min_row=ws_1.min_row,max_row=ws_1.min_row):
        if row != None:
            title_row += ([row],)
    for row in ws_2.iter_rows(min_row=ws_2.min_row,max_row=ws_2.min_row):
        if row != None:
            title_row += ([row],)
    NEW_FILE_VAlUES += (title_row,)
    
    #COMPARISON LABARATORY!
    for index, i in enumerate(ws_1.iter_rows()):
            if i[med_col_1].value != None: #medicine cell value
                cnt_same = 0
                continue_loop = False

                _value_1 = str(i[med_col_1].value).lower()
                _com_1 = str(i[com_col_1].value).lower()
                
                for new_index, row in enumerate(NEW_FILE_VAlUES[1:]):
                    for value in row[0]:
                        _value_01 = str(value[med_col_1].value).lower()
                        _com_01  = str(value[com_col_1].value).lower()
                        if _value_1.isascii() and _value_01.isascii() or not _value_1.isascii() and not _value_01.isascii(): #to check value is latin
                            pass
                        elif _value_1.isascii():
                            _value_1 = to_cyrillic(_value_1)
                        elif _value_01.isascii():
                            _value_01 = to_cyrillic(_value_01)

                        if _com_1.isascii() and _com_01.isascii() or not _com_1.isascii() and not _com_01.isascii():
                            pass
                        elif not _com_1.isascii():
                            _com_1 = to_latin(_com_1)
                        elif not _com_01.isascii():
                            _com_01 = to_latin(_com_01)

                        if _value_1 == _value_01 and _com_1 == _com_01:
                            NEW_FILE_VAlUES[new_index+1][0].append(i)
                            continue_loop = True
                            break

                        if _value_1[0:7] == _value_01[0:7]:
                            if fuzz.ratio(_com_1, _com_01) > 65:
                                if set(digit_regex(_value_1)) == set(digit_regex(_value_01)):
                                    for typ3 in TYPES:
                                        if typ3 in _value_1 and typ3 in _value_01:
                                            NEW_FILE_VAlUES[new_index+1][0].append(i)
                                            # print(_value_1, _value_01, ">>", typ3)
                                            continue_loop = True
                                            break
                                    if continue_loop:
                                        break
                                    if not any(t1p3 in _value_1 for t1p3 in TYPES) and not any(t1p3 in _value_01 for t1p3 in TYPES) or not any(t1p3 in _value_1 for t1p3 in TYPES) and  any(t1p3 in _value_01 for t1p3 in TYPES) or any(t1p3 in _value_1 for t1p3 in TYPES) and not any(t1p3 in _value_01 for t1p3 in TYPES):
                                        NEW_FILE_VAlUES[new_index+1][0].append(i)
                                        continue_loop = True
                                        break
                    if continue_loop:
                        break
                if continue_loop:
                    continue
                print(index)
                for j in ws_2.iter_rows():          #to iter second col
                    if j[med_col_2].value != None:
                        _value_2 = str(j[med_col_2].value).lower()
                        _com_2 = str(j[com_col_2].value).lower()
                        if med_col_2 < len(j):

                                if _value_1.isascii() and _value_2.isascii() or not  _value_1.isascii() and  not _value_2.isascii(): #to check value is latin
                                    pass
                                elif _value_1.isascii():
                                    _value_1 = to_cyrillic(_value_1)
                                elif _value_2.isascii():
                                    _value_2 = to_cyrillic(_value_2)
                             

                                if _com_1.isascii() and _com_2.isascii() or not _com_1.isascii() and not _com_2.isascii():
                                    pass
                                elif not _com_1.isascii():
                                    _com_1 = to_latin(_com_1)
                                elif not _com_2.isascii():
                                    _com_2 = to_latin(_com_2)
                                
                                if _value_1 == _value_2 and _com_1 == _com_2:
                                    if cnt_same == 0:
                                        NEW_FILE_VAlUES += (([i],[j]),)
                                    else:
                                        NEW_FILE_VAlUES[-1][1].append(j)
                                    cnt_same += 1
                                    continue

                                if  _value_1[0:7] == _value_2[0:7]:
                                    if fuzz.ratio(_com_1[0:6], _com_2[0:6]) > 65 or fuzz.ratio(_com_1, _com_2) > 65:
                                        
                                        for typ3 in TYPES:
                                            if typ3 in _value_1 and typ3 in _value_2:
                                                
                                                for measure in MEASURES:
                                                    if measure in _value_1 and measure in _value_2:
                                                        if set(digit_regex(_value_1)) == set(digit_regex(_value_2)):
                                                            if cnt_same == 0:
                                                                NEW_FILE_VAlUES += (([i],[j]),)
                                                            else:
                                                                NEW_FILE_VAlUES[-1][1].append(j)
                                                            cnt_same += 1
                                                            break
                                                        break

                                                if not any(m3asure in _value_1 for m3asure in MEASURES) and not any(m3asure in _value_2 for m3asure in MEASURES):
                                                    if set(digit_regex(_value_1)) == set(digit_regex(_value_2)):          
                                                        if cnt_same == 0:
                                                            NEW_FILE_VAlUES += (([i],[j]),)
                                                        else:
                                                            NEW_FILE_VAlUES[-1][1].append(j)
                                                        cnt_same += 1
                                                        break
                                                    break

                                        if not any(t1p3 in _value_1 for t1p3 in TYPES) and not any(t1p3 in _value_2 for t1p3 in TYPES) or not any(t1p3 in _value_1 for t1p3 in TYPES) and any(t1p3 in _value_2 for t1p3 in TYPES) or any(t1p3 in _value_1 for t1p3 in TYPES) and not any(t1p3 in _value_2 for t1p3 in TYPES):
                                        
                                                if set(digit_regex(_value_1)) == set(digit_regex(_value_2)):          
                                                    if cnt_same == 0:
                                                        NEW_FILE_VAlUES += (([i],[j]),)
                                                    else:
                                                        NEW_FILE_VAlUES[-1][1].append(j)
                                                    cnt_same += 1
                                                    continue
                if cnt_same == 0 and index != 0:
                    NEW_FILE_VAlUES += (([i],[("NO",)]),)

    wb.close()
    create_excel(NEW_FILE_VAlUES)

def create_excel(values):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "result"
    cnt_col = 1
    for index, data in enumerate(values):

        for zero_values in  data[0]:
            cnt_same_str = 5
            for col_index,  cell in enumerate(zero_values):
                if cell.value != None:
                    if index == 0:
                        cnt_col += 1
                    if sheet.cell(row=index+1, column=col_index+1).value == None:
                        sheet.cell(row=index+1, column=col_index+1).value = cell.value
                    else:
                        sheet.cell(row=index+1, column=col_index+1).value = str(sheet.cell(row=index+1, column=col_index+1).value)
                        sheet.cell(row=index+1, column=col_index+1).value += "\n" + str(cell.value)
                        sheet.row_dimensions[index+1].height = cnt_same_str * 10 #height of rows (*10 mm > sm)
                        cnt_same_str += 1
                    if type(cell.value) == str and index != 0:
                        sheet.column_dimensions[get_column_letter(col_index+1)].width = 30 #this is width of the columns
                
        for first_values in data[1]:
            for col_index, cell in enumerate(first_values):
                if type(cell) != str:
                    if cell.value != None:
                        if sheet.cell(row=index+1, column=cnt_col+col_index+1).value == None:
                            sheet.cell(row=index+1, column=cnt_col+col_index+1).value = cell.value
                        else:
                            sheet.cell(row=index+1, column=cnt_col+col_index+1).value = str(sheet.cell(row=index+1, column=cnt_col+col_index+1).value)
                            sheet.cell(row=index+1, column=cnt_col+col_index+1).value += "\n" + str(cell.value)
                            sheet.row_dimensions[index+1].height = cnt_same_str * 10 #height of rows (*10 mm > sm)
                            cnt_same_str += 1
                        if type(cell.value) == str and index != 0:
                            sheet.column_dimensions[get_column_letter(cnt_col+col_index+1)].width = 30 #this is width of the columns
                        
                elif type(cell) == str:
                    sheet.cell(row=index+1, column=cnt_col+col_index+3).value = str(cell)
    
    wb.save(MEDIA_ROOT / "sample.xlsx")
    wb.close()